import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

# Load the Excel file
file_path = 'MM_FAIRData.xlsx'
sheet_name = 'Overall Results'  # Sheet name containing the data
df = pd.read_excel(file_path, sheet_name=sheet_name, header=4)  # Assuming data starts from row 5 (0-indexed)

# Plotting FAIR parameters
if 'FAIR' in df.columns:
    fair_parameters = ['FINDABLE', 'ACCESSIBLE', 'INTEROPERABLE', 'REUSABLE']
    data_to_plot = df[df['FAIR'].isin(fair_parameters)]

    plt.figure(figsize=(10, 6))
    for parameter in fair_parameters:
        parameter_data = data_to_plot[data_to_plot['FAIR'] == parameter]
        plt.bar(parameter, parameter_data['Overall'].iloc[0], label=parameter, alpha=0.7)  # Bar plot

    plt.xlabel('FAIR Parameters')
    plt.ylabel('Overall Amounts')
    plt.title('Comparison of FAIR Parameters')
    plt.legend()
    plt.grid(axis='y')  # Grid on y-axis for bar plot
    plt.tight_layout()
    plt.show()

    # Print non-empty values from specified range
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=14, max_row=19, min_col=1, max_col=2, values_only=True):
        non_empty_values = [cell for cell in row if cell is not None]
        if non_empty_values:
            print(non_empty_values)

    # Create TTL file with extracted data
    levels = []
    for row in sheet.iter_rows(min_row=14, max_row=19, min_col=1, max_col=1, values_only=True):
        levels.extend(cell[0] for cell in row if cell is not None and cell[0] is not None)

    comments = []
    for row in sheet.iter_rows(min_row=14, max_row=19, min_col=2, max_col=2, values_only=True):
        comments.extend(cell[0] for cell in row if cell is not None and cell[0] is not None)

    ttl_file_path = 'output.ttl'
    with open(ttl_file_path, 'w') as ttl_file:
        for index, row in df.iterrows():
            ttl_file.write(f'<FAIR>{row["FAIR"]}\n')
            ttl_file.write(f'\t<Overall>{row["Overall"]}\n')
            ttl_file.write('.\n\n')

        for level, comment in zip(levels, comments):
            ttl_file.write(f'<Level>{level}\n')
            ttl_file.write(f'\t<HasComment>{comment}\n')
            ttl_file.write('.\n\n')
